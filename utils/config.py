from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from colorama import (
    Fore,
    Back,
    Style
)

import argparse
import time
import os
from constants import (
    EXCEL_FILE_NAME,
    DOWNLOAD_FOLDERS,
    EXCEL_SHEET_NAME
)


class ConfigBuilder:
    def __init__(self) -> None:
        self.n_rows = 0
        self.n_cols = 0

    def new_workbook(self) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = EXCEL_SHEET_NAME
        return wb

    def add_config_header(self, ws: Workbook) -> None:
        headers = ["Folder Name", "Extension"]
        self.n_cols = len(headers)
        ws.append(headers)

    def add_config_rows(self, ws: Workbook) -> None:
        for folder_name, extensions in DOWNLOAD_FOLDERS.items():
            for extension in extensions:
                ws.append([folder_name, extension])
                self.n_rows += 1

    def get_ref_table(self, ws: Workbook) -> str:
        return f"A1:{chr(65 + self.n_cols - 1)}{self.n_rows + 1}"

    def create_table(self, ws: Workbook) -> None:
        ref = self.get_ref_table(ws)
        tab = Table(displayName=f"{EXCEL_SHEET_NAME}Table", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    def save_table(self, wb: Workbook) -> None:
        print(f"Saving config file: {EXCEL_FILE_NAME}")
        wb.save(EXCEL_FILE_NAME)

    def create_config_file(self) -> None:
        wb: Workbook = self.new_workbook()
        ws = wb["Config"]

        self.add_config_header(ws)
        self.add_config_rows(ws)
        self.create_table(ws)
        self.save_table(wb)


class Config(ConfigBuilder):
    def __init__(self, args: argparse.Namespace) -> None:
        super().__init__()
        self.config: dict = {}
        self.args: argparse.Namespace = args

    def does_config_file_exist(self) -> bool:
        return os.path.exists(os.path.join(os.getcwd(), EXCEL_FILE_NAME))

    def fetch_config(self) -> dict:
        if not self.does_config_file_exist():
            response: str = "y"

            if not self.args.run_non_interactive:
                response: str = input(
                    f"{Fore.YELLOW }I see you don't have a configuration file yet. \n"
                    f"{Style.RESET_ALL}It will be an Excel file called {Fore.GREEN}'{EXCEL_FILE_NAME}'{Style.RESET_ALL}. \n"
                    "This is where you can specify the folders and extensions you want to organize. \n"
                    f"Do you want to create it? {Fore.GREEN}(y/n): "
                )

            if response.lower() == "y":
                print("\nCreating the Excel file for you with some default configurations...\n")
                self.create_config_file()

                if not self.args.run_non_interactive:
                    input(
                        f"\n{Fore.GREEN}Hooray!{Style.RESET_ALL} I've created a file called {Fore.GREEN}'{EXCEL_FILE_NAME}'{Style.RESET_ALL} for you! I placed it in the same directory as this script. \n"
                        "Please take some time to review the file. I've added some defaults for you. \n"
                        f"{Fore.GREEN}Press enter to exit the program...\n"
                    )
                exit()
            else:
                print(f"{Fore.RED}Awhh, okay! You didn't say 'y' for yes. I'll exit now. Goodbye!")
                print(":'(")
                time.sleep(2)
                exit()
        else:
            print(f"{Style.RESET_ALL}Found the {Fore.GREEN}'{EXCEL_FILE_NAME}'{Style.RESET_ALL} file. Reading config...")
            self.read_config_file()
        return self.config

    def clean_config_file(self) -> None:
        for folder_name, extensions in self.config.items():
            extensions = [ext.lower() for ext in extensions]
            self.config[folder_name] = list(set(extensions))

    def read_config_file(self) -> dict:
        wb = load_workbook(EXCEL_FILE_NAME)
        try:
            ws = wb[EXCEL_SHEET_NAME]
        except KeyError:
            print(f"{Fore.RED}Something went wrong! The '{EXCEL_SHEET_NAME}' sheet is missing. Please check the Excel file and try again.")

            if not self.args.run_non_interactive:
                input(f"{Fore.GREEN}Press enter to exit...")
            exit()

        for row in ws.iter_rows(min_row=2, values_only=True):
            folder_name, extension = row
            if folder_name not in self.config:
                self.config[folder_name] = [extension]
            else:
                self.config[folder_name].append(extension)

        self.clean_config_file()
        return self.config

    def summarize_config(self) -> None:
        # Print a table where the first column is the folder name and the second is the number of extensions
        print("\nHere is a summary fo you of all the file extensions I found!")
        print(f"{'Folder Name':<20} | {'Number of Extensions':<20}")
        print("-" * 40)
        for folder_name, extensions in self.config.items():
            print(f"{folder_name:<20} | {len(extensions):<25}")
        print("\n")

        if not self.args.run_non_interactive:
            input(f"{Fore.GREEN}Press enter to continue...")
